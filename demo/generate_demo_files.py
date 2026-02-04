"""Generate demo Excel files for stock and sales data."""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Stock data - all plants with inventory
stock_data = [
    ("Monstera Deliciosa", 45, 29.99),
    ("Fiddle Leaf Fig", 30, 49.99),
    ("Snake Plant", 80, 15.99),
    ("Pothos Golden", 120, 12.99),
    ("Peace Lily", 55, 18.99),
    ("Rubber Plant", 40, 24.99),
    ("Bird of Paradise", 25, 59.99),
    ("ZZ Plant", 65, 22.99),
    ("Spider Plant", 90, 9.99),
    ("Philodendron Brasil", 70, 14.99),
    ("Calathea Orbifolia", 35, 34.99),
    ("Aloe Vera", 100, 11.99),
]

# Sales data - subset of plants with some quantities exceeding stock
sales_data = [
    ("Monstera Deliciosa", 38),      # Stock: 45, OK
    ("Fiddle Leaf Fig", 42),          # Stock: 30, EXCEEDS by 12
    ("Snake Plant", 75),              # Stock: 80, OK
    ("Pothos Golden", 150),           # Stock: 120, EXCEEDS by 30
    ("Peace Lily", 55),               # Stock: 55, EXACT
    ("Bird of Paradise", 30),         # Stock: 25, EXCEEDS by 5
    ("ZZ Plant", 60),                 # Stock: 65, OK
    ("Calathea Orbifolia", 50),       # Stock: 35, EXCEEDS by 15
    ("Aloe Vera", 85),                # Stock: 100, OK
]

def create_styled_header(ws, headers):
    """Apply styling to header row."""
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

def create_stock_file():
    """Create the stock inventory Excel file."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventory"
    
    # Headers
    headers = ["Product Name", "Quantity In Stock", "Price Per Unit ($)"]
    create_styled_header(ws, headers)
    
    # Data
    for row, (name, qty, price) in enumerate(stock_data, 2):
        ws.cell(row=row, column=1, value=name)
        ws.cell(row=row, column=2, value=qty)
        ws.cell(row=row, column=3, value=price)
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    
    wb.save("stock_inventory.xlsx")
    print("Created: stock_inventory.xlsx")

def create_sales_file():
    """Create the sales data Excel file."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales"
    
    # Headers
    headers = ["Product Name", "Quantity Sold"]
    create_styled_header(ws, headers)
    
    # Data
    for row, (name, qty) in enumerate(sales_data, 2):
        ws.cell(row=row, column=1, value=name)
        ws.cell(row=row, column=2, value=qty)
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 18
    
    wb.save("sales_data.xlsx")
    print("Created: sales_data.xlsx")

if __name__ == "__main__":
    create_stock_file()
    create_sales_file()
    print("\nDemo files generated successfully!")
    print("\nStock file contains 12 plants with quantities and prices.")
    print("Sales file contains 9 plants (subset) with some quantities exceeding stock:")
    print("  - Fiddle Leaf Fig: sold 42, stock 30 (exceeds by 12)")
    print("  - Pothos Golden: sold 150, stock 120 (exceeds by 30)")
    print("  - Bird of Paradise: sold 30, stock 25 (exceeds by 5)")
    print("  - Calathea Orbifolia: sold 50, stock 35 (exceeds by 15)")
