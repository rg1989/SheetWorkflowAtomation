"""
Export utilities for Excel and PDF.
"""
import pandas as pd
from typing import List, Dict, Any
import os
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

from app.models.diff import CellChange, ChangeType


class ExcelExporter:
    """Export DataFrames to Excel with formatting."""
    
    def export(
        self,
        df: pd.DataFrame,
        output_path: str,
        changes: List[CellChange] = None,
        highlight_changes: bool = True,
    ) -> str:
        """
        Export a DataFrame to Excel with optional change highlighting.
        
        Args:
            df: DataFrame to export
            output_path: Path for output file
            changes: List of changes to highlight
            highlight_changes: Whether to highlight changed cells
            
        Returns:
            Path to the output file
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4A5568", end_color="4A5568", fill_type="solid")
        modified_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
        border = Border(
            left=Side(style='thin', color='E2E8F0'),
            right=Side(style='thin', color='E2E8F0'),
            top=Side(style='thin', color='E2E8F0'),
            bottom=Side(style='thin', color='E2E8F0'),
        )
        
        # Build set of changed cells for quick lookup
        changed_cells = set()
        if changes and highlight_changes:
            for change in changes:
                changed_cells.add((change.row, change.column))
        
        # Write data
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True)):
            for c_idx, value in enumerate(row):
                cell = ws.cell(row=r_idx + 1, column=c_idx + 1, value=value)
                cell.border = border
                cell.alignment = Alignment(horizontal='left', vertical='center')
                
                if r_idx == 0:  # Header row
                    cell.font = header_font
                    cell.fill = header_fill
                else:
                    # Check if this cell was changed
                    col_name = df.columns[c_idx] if c_idx < len(df.columns) else ""
                    if (r_idx - 1, col_name) in changed_cells:  # -1 because of header
                        cell.fill = modified_fill
        
        # Auto-adjust column widths
        for col_idx, column in enumerate(df.columns):
            max_length = max(
                len(str(column)),
                df[column].astype(str).str.len().max() if len(df) > 0 else 0
            )
            ws.column_dimensions[ws.cell(row=1, column=col_idx + 1).column_letter].width = min(max_length + 2, 50)
        
        # Freeze header row
        ws.freeze_panes = 'A2'
        
        # Save
        wb.save(output_path)
        return output_path


class PDFExporter:
    """Export diff summaries to PDF."""
    
    def export(
        self,
        diff_data: Dict[str, Any],
        output_path: str,
    ) -> str:
        """
        Export a diff summary to PDF.
        
        Args:
            diff_data: Diff result dictionary
            output_path: Path for output file
            
        Returns:
            Path to the output file
        """
        doc = SimpleDocTemplate(
            output_path,
            pagesize=A4,
            rightMargin=0.75*inch,
            leftMargin=0.75*inch,
            topMargin=0.75*inch,
            bottomMargin=0.75*inch,
        )
        
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=20,
        )
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=14,
            spaceAfter=10,
            spaceBefore=15,
        )
        
        elements = []
        
        # Title
        elements.append(Paragraph("Workflow Execution Summary", title_style))
        elements.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", styles['Normal']))
        elements.append(Spacer(1, 20))
        
        # Summary section
        summary = diff_data.get("summary", {})
        elements.append(Paragraph("Summary", heading_style))
        
        summary_data = [
            ["Metric", "Value"],
            ["Rows Affected", str(summary.get("rowsAffected", 0))],
            ["Cells Modified", str(summary.get("cellsModified", 0))],
            ["Total Rows", str(summary.get("totalRows", 0))],
            ["Warnings", str(summary.get("warnings", 0))],
            ["Errors", str(summary.get("errors", 0))],
        ]
        
        summary_table = Table(summary_data, colWidths=[2.5*inch, 2*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.Color(0.29, 0.33, 0.41)),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('TOPPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.Color(0.97, 0.97, 0.97)),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.Color(0.8, 0.8, 0.8)),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
            ('TOPPADDING', (0, 1), (-1, -1), 8),
        ]))
        elements.append(summary_table)
        elements.append(Spacer(1, 20))
        
        # Changes section
        changes = diff_data.get("changes", [])
        if changes:
            elements.append(Paragraph("Changes Detail", heading_style))
            
            change_data = [["Key", "Column", "Old Value", "New Value"]]
            for row_change in changes[:50]:  # Limit to first 50 rows
                for cell in row_change.get("cells", []):
                    change_data.append([
                        str(row_change.get("keyValue", "")),
                        str(cell.get("column", "")),
                        str(cell.get("oldValue", "")),
                        str(cell.get("newValue", "")),
                    ])
            
            change_table = Table(change_data, colWidths=[1.5*inch, 1.5*inch, 1.5*inch, 1.5*inch])
            change_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.Color(0.29, 0.33, 0.41)),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
                ('TOPPADDING', (0, 0), (-1, 0), 10),
                ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.Color(0.8, 0.8, 0.8)),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
                ('TOPPADDING', (0, 1), (-1, -1), 6),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.Color(0.97, 0.97, 0.97)]),
            ]))
            elements.append(change_table)
            
            if len(changes) > 50:
                elements.append(Spacer(1, 10))
                elements.append(Paragraph(
                    f"... and {len(changes) - 50} more rows",
                    styles['Normal']
                ))
        
        # Warnings section
        warnings = diff_data.get("warnings", [])
        if warnings:
            elements.append(Spacer(1, 20))
            elements.append(Paragraph("Warnings", heading_style))
            
            for warning in warnings:
                elements.append(Paragraph(
                    f"• {warning.get('message', '')}",
                    styles['Normal']
                ))
        
        # Build PDF
        doc.build(elements)
        return output_path
